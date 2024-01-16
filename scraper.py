import requests
from bs4 import BeautifulSoup
import markdown2
from openpyxl import load_workbook, Workbook
import os
import pandas as pd

def reverse_rows_with_same_date(scraped_data):
    # Convert the list of data rows to a DataFrame
    df = pd.DataFrame(scraped_data, columns=["Company", "Role", "Location", "Application/Link", "Date Posted"])

    # Ensure the DataFrame has at least 5 columns
    if df.shape[1] < 5:
        print("The DataFrame does not have enough columns.")
        return None

    # Identify groups of rows with the same date in the fifth column and reverse them
    reversed_df = pd.DataFrame()
    for date, group in df.groupby('Date Posted'):
        print(f"Reversing rows with date: {date}")
        print(group)
        reversed_group = group.iloc[::-1]
        reversed_df = pd.concat([reversed_df, reversed_group])

    # Reset index for the new DataFrame
    reversed_df = reversed_df.reset_index(drop=True)

    return reversed_df

def extract_and_write_to_xlsx(xlsx_filename):
    readme_url = "https://github.com/SimplifyJobs/Summer2024-Internships/blob/dev/README.md"
    response = requests.get(readme_url)

    if response.status_code == 200:
        html_content = markdown2.markdown(response.text)
        soup = BeautifulSoup(html_content, 'html.parser')
        table_rows = soup.find_all('tr')

        all_data = []  # List to store all rows

        for row in table_rows:
            table_cells = row.find_all('td')
            if len(table_cells) > 4 and table_cells[4].get_text(strip=True) == 'Nov 30':
                break  # Stop if 'Nov 30' is found in the 4th column
            row_data = []
            for index, cell in enumerate(table_cells):
                if index == 3:  # if this is the fourth cell (index 3)
                    link = cell.find('a')
                    row_data.append(link['href'][2:] if link else '')  # append href if link exists, else empty string
                else:
                    row_data.append(cell.get_text(strip=True))
            all_data.append(row_data)

        # Process the scraped data
        all_data.reverse()  # Reverse the list of rows
        modified_data = reverse_rows_with_same_date(all_data)
        if modified_data is None:
            return

        # Check if the Excel file exists and has content
        workbook_exists = os.path.exists(xlsx_filename)
        workbook = load_workbook(xlsx_filename) if workbook_exists else Workbook()
        sheet = workbook.active

        if not workbook_exists or sheet.max_row == 1:  # If the workbook is new or the sheet is empty
            sheet.append(["Company", "Role", "Location", "Application/Link", "Date Posted"])

        # Append data to the sheet
        for index, row in modified_data.iterrows():
            sheet.append(row.tolist())

        workbook.save(xlsx_filename)
        print(f"Data appended to {xlsx_filename}")
    else:
        print(f"Failed to retrieve the README file. Status code: {response.status_code}")

# Example usage
extract_and_write_to_xlsx('output.xlsx')





